import time
import threading
import winsound
import queue
import cv2
import os
from filelock import FileLock
import mysql.connector
import pickle
import numpy as np
from keras_facenet import FaceNet
import openpyxl
from datetime import datetime

conn = mysql.connector.connect(
    host="localhost",
    user="root",
    password="Anshul@2006",
    database="AMS" 
)
cursor = conn.cursor()

CAMERA = 0
users_path = "RegisteredUsers"
main_embeddings = "main_embeddings.pkl"

embedder = FaceNet()
totalimg = 10
names = []

embeddings_lock = threading.Lock()
recompiling = False

class CameraStream:
    def __init__(self, src=0):
        self.cap = cv2.VideoCapture(src, cv2.CAP_DSHOW)
        self.cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*"MJPG"))
        self.cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)

        self.q = queue.Queue(maxsize=1)
        self.running = False

    def start(self):
        self.running = True
        threading.Thread(target=self._capture_loop, daemon=True).start()
        return self

    def _capture_loop(self):
        while self.running:
            ret, frame = self.cap.read()
            if not ret:
                break
            if not self.q.empty():
                try: 
                    self.q.get_nowait()  
                except queue.Empty:
                    pass
            self.q.put(frame)
        self.cap.release()

    def read(self):
        return self.q.get()

    def stop(self):
        self.running = False

def mk_embeddings():
    global recompiling
    all_embedd = {}
    
    cursor.execute("SELECT userID FROM attendance_records")
    valid_ids = {str(row[0]) for row in cursor.fetchall()}
    
    for user_id in os.listdir(users_path):
        user_dir = os.path.join(users_path, user_id)
        
        if not os.path.isdir(user_dir) or user_id not in valid_ids:
            continue
            
        embed_path = os.path.join(user_dir, "embeddings.pkl")
        if os.path.exists(embed_path):
            try:
                with open(embed_path, 'rb') as f:
                    all_embedd[user_id] = pickle.load(f)
            except (pickle.PickleError, EOFError) as e:
                print(f"Error loading {embed_path}: {e}")
                continue
    lock = FileLock(f"{main_embeddings}.lock")
    with lock:
        with open(main_embeddings, 'wb') as f:
            pickle.dump(all_embedd, f)
    
    recompiling = False

def compile_embeddings(wait=False):
    global recompiling
    
    with embeddings_lock:
        if recompiling:
            return
            
        recompiling = True
        
        if wait:
            mk_embeddings()
        else:
            threading.Thread(target=mk_embeddings, daemon=True).start()

def testing():
# if the deletion of the student is in progress #
    while recompiling:
        print("Waiting for embeddings recompilation to finish...")
        time.sleep(0.5)

    if not os.path.exists(main_embeddings):
        print("Register at least 1 user to continue...")
        return
        
    with FileLock(f"{main_embeddings}.lock"):
        with open(main_embeddings, 'rb') as f:
            embeddings = pickle.load(f)
    
    last_seen = [0]  
    label_timeout = 2 
    cap = CameraStream().start()
    frame_queue = queue.Queue(maxsize=1)
    label = "Detecting..."
    face_box = None  
    stop_flag = threading.Event()

    def process_frames():
        nonlocal label, face_box
        while not stop_flag.is_set():
            try:
                frame = frame_queue.get(timeout=1)
            except queue.Empty:
                continue

            detections = embedder.extract(frame, threshold=0.8)
            if detections:
                for det in detections:
                    x, y, w, h = det['box']
                    x = max(0, x)
                    y = max(0, y)
                    face_img = frame[y:y+h, x:x+w]
                    if face_img.size == 0:
                        continue
                    face_emb = det['embedding']
                    min_dist = float("inf")
                    identity = None
                    for uid, db_emb in embeddings.items():
                        dist = np.linalg.norm(face_emb - db_emb)
                        if dist < min_dist:
                            min_dist = dist
                            identity = uid
                    if min_dist < 1.0:
                        name_st, div_st, rollno_st = get_student(identity)
                        label = f"{name_st} {div_st} ({rollno_st})"
                        face_box = (x, y, w, h) 
                        mark_attendance(name_st, div_st, rollno_st)
                    else:
                        label = "Unknown"
                        face_box = (x, y, w, h) 
            else:
                face_box = None  
                
            last_seen[0] = time.time()

    threading.Thread(target=process_frames, daemon=True).start()

    try:
        while True:
            frame = cap.read()
            display_frame = frame.copy()

            if face_box and time.time() - last_seen[0] <= label_timeout:
                x, y, w, h = face_box
                
                square_size = max(w, h)
                center_x, center_y = x + w//2, y + h//2
                half_size = square_size // 2
                x1 = max(0, center_x - half_size)
                y1 = max(0, center_y - half_size)
                x2 = min(frame.shape[1], center_x + half_size)
                y2 = min(frame.shape[0], center_y + half_size)
                
                cv2.rectangle(display_frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                
                label_size = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.8, 2)[0]
                label_x = center_x - label_size[0] // 2
                label_y = max(20, y1 - 10)  
                
                cv2.rectangle(display_frame, 
                              (label_x - 5, label_y - label_size[1] - 5),
                              (label_x + label_size[0] + 5, label_y + 5),
                              (0, 0, 0), -1)
                
                cv2.putText(display_frame, label, (label_x, label_y),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
            else:
                label = "Place your face in the middle"
                face_box = None
                cv2.putText(display_frame, label, (10, 30),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

            cv2.imshow("FaceRecognition", display_frame)

            if not frame_queue.full():
                frame_queue.put(frame)

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
    finally:
        stop_flag.set()
        cap.stop()
        cv2.destroyAllWindows()

def getUserId():
    pathuserid = "UserIdCounter.txt"
    if not os.path.exists(pathuserid):
        with open(pathuserid, "w") as f:
            f.write("1")
        return 1
    with open(pathuserid, "r") as f:
        user_id = int(f.read())
    with open(pathuserid, "w") as f:
        f.write(str(user_id + 1))
    return user_id

def mark_attendance(name, div, rollno):
    folder = "Data"
    os.makedirs(folder, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    file_path = os.path.join(folder, f"{today}.xlsx")

    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Class", "Roll No", "Time"])
    else:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == name and row[1] == div and row[2] == rollno:
                return 

    time_str = datetime.now().strftime("%H:%M:%S")
    ws.append([name, div, rollno, time_str])
    wb.save(file_path)
    winsound.Beep(2000, 1000)

def add_students(userid, name, div, rollno):
    query = "INSERT INTO attendance_records VALUES(%s, %s, %s, %s)"
    cursor.execute(query, (userid, name, div, rollno))
    conn.commit()

def get_student(userid):
    query = "SELECT name, class, rollno FROM attendance_records where userID=%s"
    cursor.execute(query, (userid,))
    result = cursor.fetchone()
    if result:
        return result
    return ("Unknown", "Unknown", "Unknown")

def registerUsers(user_id):
    name = input("Enter your name: ")
    div = input("Enter class: ")
    rollno = int(input("Enter rollno: "))
    add_students(user_id, name, div, rollno)

    user_dir = os.path.join(users_path, str(user_id))
    os.makedirs(user_dir, exist_ok=True)

    cap = CameraStream().start()
    frame_queue = queue.Queue(maxsize=10)

    count = 0
    frame_count = 0
    stop_thread = False
    user_embeddings = []

    def process_frames():
        nonlocal count
        while not stop_thread and count < totalimg:
            if not frame_queue.empty():
                frame = frame_queue.get()
                if frame is None:
                    continue

                detections = embedder.extract(frame, threshold=0.8)
                if detections:
                    det = detections[0]
                    x, y, w, h = det['box']
                    x = max(0, x)
                    y = max(0, y)
                    face_img = frame[y:y+h, x:x+w]
                    if face_img.size > 0:
                        img_path = os.path.join(user_dir, f"{count}.jpg")
                        cv2.imwrite(img_path, face_img)
                        vec = det["embedding"]
                        user_embeddings.append(vec)
                        count += 1

    processing_thread = threading.Thread(target=process_frames)
    processing_thread.start()

    while count < totalimg:
        frame = cap.read()
        frame_count += 1

        if frame_count % 3 == 0 and not frame_queue.full():
            frame_queue.put(frame.copy())

        cv2.putText(frame, f"Registering face {count}/{totalimg}", (20, 50),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 0), 2)
        cv2.imshow("Registering", frame)
        if cv2.waitKey(1) == 27:
            break

    stop_thread = True
    processing_thread.join()
    cap.stop()
    cv2.destroyAllWindows()

    if user_embeddings:
        user_embedding_path = os.path.join(user_dir, "embeddings.pkl")
        with open(user_embedding_path, 'wb') as f:
            pickle.dump(np.mean(user_embeddings, axis=0), f)
        compile_embeddings(wait=True)

def totalEmployees():
    query = 'SELECT COUNT(userID) from attendance_records'
    cursor.execute(query)
    totalemployee = cursor.fetchone()
    print(totalemployee)

def totalPresent():
    folder = "Data"
    os.makedirs(folder, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    file_path = os.path.join(folder, f"{today}.xlsx")
    ans = 0
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        s = wb.active
        ans = s.max_row
    if ans:
        print(ans-1)
    else :
        print(ans)

def delete_data(userid):
    query = "DELETE FROM attendance_records WHERE userID=%s"
    cursor.execute(query, (userid,))
    conn.commit()

def get_details(name, div, rollno):
    try:
        query = "SELECT userID FROM attendance_records WHERE name=%s and class=%s and rollno=%s;"
        cursor.execute(query, (name, div, rollno))
        res = cursor.fetchone()
        return res[0]
    except:
        return -1

def delete_user():
    name = input("Enter name of the student to be deleted : ")
    div = input(f"Enter class of {name} : ")
    rollno = int(input(f"Enter Roll number of {name} : "))
    user_id = get_details(name, div, rollno)
    
    if user_id == -1:
        print("No user Exists with these credentials!!")
        return
        
    delete_data(user_id)
    
    user_dir = os.path.join(users_path, str(user_id))
    if os.path.exists(user_dir):
        import shutil
        shutil.rmtree(user_dir)
    compile_embeddings(wait=True)
    
    print("Successfully deleted student's data!!")

def main_menu():
    while True:
        print("\nAttendance Management System")
        print("1. Register New Student")
        print("2. Start Attendance")
        print("3. Delete Student")
        print("4. Exit")
        
        choice = input("Enter your choice: ")
        
        if choice == "1":
            user_id = getUserId()
            registerUsers(user_id)
        elif choice == "2":
            testing()   
        elif choice == "3":
            delete_user()
        elif choice == "4":
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    if not os.path.exists(main_embeddings):
        compile_embeddings(wait=True)
    
    try:
        main_menu()
    finally:
        cursor.close()
        conn.close()